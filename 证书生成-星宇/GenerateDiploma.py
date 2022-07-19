import cv2
from PIL import ImageFont, ImageDraw, Image
import numpy as np
import tkinter as tk 
from tkinter.filedialog import askdirectory,askopenfilename
import os
import openpyxl

def selectfilename():
    #选择文件
    path_ = askopenfilename()
    path1.set(path_)

def selectsavedir():
    #选择文件path_接收文件地址
    path_ = askdirectory()
    path2.set(path_)


def change_img(text = "",filename = "NoFileName.jpg"):
    bk_img = cv2.imread("证书.png")
    #设置需要显示的字体
    fontpath = "阿里巴巴普惠体M.ttf"
    font = ImageFont.truetype(fontpath, 30)
    img_pil = Image.fromarray(bk_img)
    draw = ImageDraw.Draw(img_pil)
    draw.text((255, 298), text, font = font, fill = (0, 0, 0))
    bk_img = np.array(img_pil)
    cv2.imwrite(filename,bk_img)



def run():
    varx.set("...运行中...")
    global path1,path2
    openfile = path1.get()
    savepath = path2.get()
    wb = openpyxl.load_workbook(openfile)
    ws = wb.active
    info_list = []
    for i in range(1,ws.max_row+1):
        info_list.append(ws["A"+str(i)].value)

    for item in info_list:
        change_img(text=item,filename= os.path.join(savepath,item+".jpg"))
    varx.set("运行完成～～")
    




if __name__ == "__main__":
    #主窗口程序
    main_box = tk.Tk()
    path1 = tk.StringVar()
    global varx
    varx = tk.StringVar()
    varx.set("等待运行中")

    tk.Label(main_box,text = "文件路径:").grid(row = 0, column = 0)
    tk.Entry(main_box, textvariable = path1).grid(row = 0, column = 1)
    tk.Button(main_box, text = "文件选择", command = selectfilename).grid(row = 0, column = 2)
    path2 = tk.StringVar()
    tk.Label(main_box,text = "存储路径:").grid(row = 1, column = 0)
    tk.Entry(main_box, textvariable = path2).grid(row = 1, column = 1)
    tk.Button(main_box, text = "路径选择", command = selectsavedir).grid(row = 1, column = 2)
    tk.Button(main_box,text = "运行",command = run).grid(row = 3, column = 1)
    
    tk.Label(main_box,textvariable=varx ).grid(row = 4, column = 1)
    main_box.mainloop()