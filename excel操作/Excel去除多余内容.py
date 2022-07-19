import os
import openpyxl
import re 
from openpyxl.utils import get_column_letter

files_path = "/Users/MikeImac/Desktop/Python—test/Excel模板/3.可视化图表模板/可视化图表【200个图表】/12-多图结合式"
os.chdir(files_path)
file_list = os.listdir()
reg = re.compile(r'^周达学教育|^ppt|^Excel|^http',re.S)

try :
    file_list.remove(".DS_Store")
except:
    print("No .DS_Store")

def delExcelname(pathlist):
    newlist = []
    for item in pathlist:
        if "周达学教育" in item:
            rep = item
            rep = rep.replace("周达学教育","")
            newlist.append(rep)
        else:
            newlist.append(item)
    return newlist

def del_excel(path):
    wb = openpyxl.load_workbook(path,data_only = True)
    ws = wb.active
    # ws.title = "test"
    my_list = []
    maxrows = ws.max_row  

    for i in range(maxrows):
        temp_list = []
        for each in ws.iter_cols():
            temp_list.append(each[i].value)
        my_list.append(temp_list)
    change_list_pos = []
    for item in range(len(my_list)):
        for x in range(len(my_list[item])):
            a = re.match(reg,str(my_list[item][x]))
            if isinstance(a,re.Match):
                change_list_pos.append([item+1,x+1])
    for x in change_list_pos:
        x[1] = get_column_letter(x[1])
    for y in change_list_pos:
        a = ws[y[1]+str(y[0])]
        a.value = None
    return wb

a = delExcelname(file_list)
if not os.path.exists(files_path+"/new"):
    os.mkdir(files_path+"/new")
    fi =open(files_path+"/new"+"/log.txt","w") 

os.chdir(files_path+"/new")
flag = 0 
for i in range(len(file_list)):
    flag+=1
    try :
        print("Working:" + file_list[i])
        wba = del_excel(files_path + "/" +file_list[i])
        wba.save(files_path+"/new/"+a[i])
    except:
        fi.write(file_list[i]+"\n")
    finally:
        print("Done:"+file_list[i])

fi.close()