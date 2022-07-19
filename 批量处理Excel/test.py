#导入相关库
import openpyxl

#加载文件
wb  = openpyxl.load_workbook("/Users/MikeImac/Downloads/各类代码/批量处理Excel/批量处理练习表格.xlsx",data_only = False)

#获取工作表
ws = wb["销项发票"]

#获取最大行
maxrow = ws.max_row

#创建辅助信息集
Customdata = set()
datalist = {}
texislist = {}

#遍历有效数据集
for x in range(5,maxrow):
    #单行信息集
    item = {} 
    
    #是否收集了公司信息集。
    if ws["E"+str(x)].value not in Customdata:
        #发票代码收集。
        texislist[ws["E"+str(x)].value] = ws["C"+str(x)].value
        
        #添加到集合，收集公司信息。
        Customdata.add(ws["E"+str(x)].value)
    
    #判断公司字典是否已经存在键，无则新建
    if ws["E"+str(x)].value not in datalist.keys():
        datalist[ws["E"+str(x)].value] = []
    
    #将对应的信息，以字典形式进行整理，并添加到对应键的列表中。
    item["货物名称"] = ws["G"+str(x)].value
    item["规格型号"] = ws["H"+str(x)].value
    item["单位"] = ws["I"+str(x)].value
    item["数量"] = ws["J"+str(x)].value
    item["单价"] = ws["K"+str(x)].value
    item["金额"] = ws["O"+str(x)].value
    item["日期"] = ws["D"+str(x)].value
    
    #将单行信息添加到对应的公司键下的字典中。
    datalist[ws["E"+str(x)].value].append(item)

#遍历所用公司。    
for x in Customdata:
    #复制并创建对应的表
    new_ws = wb.copy_worksheet(wb["Sheet1"])
    
    #设置Title
    new_ws.title = x
    
    #设置公司名字
    new_ws["B5"].value = x 
    
    #添加日期
    new_ws["F5"].value = datalist[x][0]["日期"]
    
    #添加经办人
    new_ws["E22"].value = "小红"
    
    #添加发票号码。
    new_ws["F6"].value = texislist[x]
    
    #创建计算变量，计算总价格
    moneycount = 0 
    
    #遍历该公司下的所有信息，形式为字典列表。
    for y in range(len(datalist[x])):
        #计算总价格。
        moneycount += int(datalist[x][y]["金额"])
        
        #添加对应消息到对应位置，第一行信息为9，因此需要加上9
        new_ws["A" + str(9+y)].value = datalist[x][y]["货物名称"]
        new_ws["B" + str(9+y)].value = str(datalist[x][y]["规格型号"])
        new_ws["C" + str(9+y)].value = str(datalist[x][y]["单位"])
        new_ws["D" + str(9+y)].value = int(datalist[x][y]["数量"])
        new_ws["E" + str(9+y)].value = int(datalist[x][y]["单价"])
        new_ws["F" + str(9+y)].value = int(datalist[x][y]["金额"])
    
    #添加对应的总价格到目标位置。
    new_ws["F21"].value = moneycount

    #添加数字转中文大学的Excel函数公式。
    new_ws["C21"].value = '=IF(F21=0,"",IF(F21<0,"负","")&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(TEXT(INT(ABS(F21)),"[DBNum2]")&"元"&TEXT(RIGHT(TEXT(F21,".00"),2),"[DBNum2]0角0分"),"零角零分","整"),"零分","整"),"零角","零"),"零元零",""))'

#保存为新的Excel
wb.save("xxxx.xlsx")
