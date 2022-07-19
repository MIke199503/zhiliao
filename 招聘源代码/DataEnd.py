import io
import msoffcrypto
import openpyxl
import time

class DataDeal:
    '''
    it's a object for organize data  and let data return..
    '''

    def __init__(self,filePath) :
        '''
        get password down!
        '''
        decrypted_workbook = io.BytesIO()
        with open(filePath, 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password='88664466')
            office_file.decrypt(decrypted_workbook)

        # data_only  mean : ignore the formula
        self.wb = openpyxl.load_workbook(decrypted_workbook,data_only=True)
        
        self.ws = self.wb["岗位JD信息"]

        #if you want to get right locationlist ,please use generateData function first 
        self.locationlist  = set() 

        self.resource = {} 



    def generateData(self):
        '''
        a function to get all the Job Detail back

        JD =  {
            "location1":{
               "JbID":{
                "平台":self.ws["B"+str(x)].value,
                "岗位":self.ws["D"+str(x)].value,
                "公司":self.ws["E"+str(x)].value,
                "工作描述":self.ws["F"+str(x)].value,
                "任职需求":self.ws["G"+str(x)].value,
                "薪资待遇":self.ws["E"+str(x)].value,
                "直达链接":self.ws["E"+str(x)].value,
               }
            }
        }
        '''
        #init a dict for save data
        JDDICT = {}
        
        #Auxiliary information
        rowMax = self.ws.max_row
        for x in range(3,rowMax):
            location = self.ws["C"+str(x)].value

            #bad Data
            if location == None:
                continue

            #clear Data & add data into __JDDICT__
            location  =  location.strip()
            if location not in JDDICT.keys() :
                JDDICT[location] = {}

            self.locationlist.add(location)
            
            #generate the data 
            JDDICT[location][self.ws["A"+str(x)].value] = {
                "平台":self.ws["B"+str(x)].value,
                "岗位":self.ws["D"+str(x)].value,
                "公司":self.ws["E"+str(x)].value,
                "工作描述":self.ws["F"+str(x)].value,
                "任职需求":self.ws["G"+str(x)].value,
                "薪资待遇":self.ws["H"+str(x)].value,
                "直达链接":self.ws["I"+str(x)].value,
            }
        self.resource = JDDICT

        return JDDICT


    def getItem(self,plats:str,loca:str):
        '''
        return Data which is suit for param : plats,loca
        '''
        data = []

        #No location
        if loca not in self.resource.keys():
            return "暂无相关地区数据"

        for jd in self.resource[loca].keys():
            if self.resource[loca][jd]["平台"] == plats:
                data.append(self.resource[loca][jd])
        return data if len(data) >=1 else "该平台暂无该地区相关数据"


    def getUserData(self):
        '''
        get all the account data , it's organized just like follow:
        {
            username:{ 
                "password": ***,
                "quality" : ***,  #Boolean Data ,You can use this data to verify account is still work or not
                }
        }
        '''
        Today = time.localtime()

        ws = self.wb['UserData']
        maxrow = ws.max_row
        UserData  = {}
        for x in range(2,maxrow):
            userName = ws["A"+str(x)].value
            userPassword = ws["B"+str(x)].value
            if userName!= None and userPassword != None:
                EndTime = time.strptime(str(ws["E"+str(x)].value),"%Y-%m-%d %H:%M:%S")
                NotOverTime = True if Today <= EndTime else False
                UserData[userName] = {
                    'password': userPassword,
                    'quality' : NotOverTime,
                }
            
        return UserData

    


if __name__ == "__main__":
    PATHS = '/Users/MikeImac/Desktop/招聘/财会岗位JD搜索.xlsx'
    dataa  = DataDeal(filePath=PATHS)
    data = dataa.generateData()
    datd1=  dataa.getUserData()
    print(datd1)

